"""Amazon category template generator.

Builds an .xlsx that mimics an Amazon flat-file / category template and fills it
with the AI-generated fields plus some default values. Swap the columns and
defaults for your real Amazon category template later.
"""
from typing import List, Dict, Optional, Union
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# --- Default values that pre-fill on every row ---
# TODO: replace with the real defaults for your Amazon category.
DEFAULT_VALUES = {
    "brand": "MY_BRAND",
    "manufacturer": "MY_BRAND",
    "condition": "New",
    "currency": "USD",
    "fulfillment": "FBA",
}

# Column order of the generated template.
COLUMNS = [
    "sku",
    "title",
    "description",
    "bullet_point_1",
    "bullet_point_2",
    "bullet_point_3",
    "keywords",
    "brand",
    "manufacturer",
    "condition",
    "currency",
    "fulfillment",
    "image_name",
]


# Which template each store country maps to. Placeholder — swap for your real
# per-country Amazon templates later.
COUNTRY_TEMPLATES = {
    "UAE": "amazon_template_uae",
    "India": "amazon_template_india",
}


def template_for_country(country: str) -> str:
    return COUNTRY_TEMPLATES.get(country, "amazon_template_default")


def generate_amazon_template(rows: List[dict], output_dir: str, batch_name: str, country: str = "") -> str:
    """Create an Amazon template xlsx from processed image rows.

    `rows` is a list of the dicts returned by ai_stub.analyze_image, each also
    carrying an "image_name" key.
    Returns the path to the generated file.
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    for i, data in enumerate(rows, start=1):
        bullets = data.get("bullet_points", []) + ["", "", ""]
        record = {
            "sku": f"{batch_name}-{i:04d}",
            "title": data.get("title", ""),
            "description": data.get("description", ""),
            "bullet_point_1": bullets[0],
            "bullet_point_2": bullets[1],
            "bullet_point_3": bullets[2],
            "keywords": ", ".join(data.get("keywords", [])),
            "image_name": data.get("image_name", ""),
            **DEFAULT_VALUES,
        }
        for col_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=i + 1, column=col_idx, value=record.get(col, ""))

    # widen a few columns for readability
    for col_idx, width in {1: 18, 2: 40, 3: 50, 7: 30}.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    template = template_for_country(country)
    ws.title = (country or "Template")[:31]

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c for c in batch_name if c.isalnum() or c in ("-", "_")) or "batch"
    prefix = template  # e.g. amazon_template_uae / _india / _default
    filename = f"{prefix}_{safe_name}_{stamp}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    return path
